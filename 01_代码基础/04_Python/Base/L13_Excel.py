# 表格
# 处理表格文件 .xlsx

import openpyxl
# 读取表格
# 这里的相对路径是cmd中的运行的路径的相对
wb = openpyxl.load_workbook('.\Python\Base\example.xlsx')
# 获取表格的所有表名
print(wb.sheetnames)
# 激活表
sheet = wb['Sheet1']
sheet = wb.active

# 获取单元格(cell)
sheet['A1'] # 获取A1单元格(cell)，里面有value、row、column、coordinate等属性，（coordinate是单元格的坐标，即A1、B1等）
# 列用字母表示，超过26列用AA、AB、AC表示
print(sheet['A1'].value)
print(sheet['B1'].value)
print(sheet['C1'].row)
print(sheet['C1'].column)
print(sheet['C1'].coordinate)
print()
# 确定表格的大小
print(sheet.max_row)
print(sheet.max_column)
print()

# 列字母和数字的转换
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(column_index_from_string('A'))
print()

# 获取行和列
# 获得矩形区域的单元格
cell_range = sheet['A1':'D4'] # 获得所有cell对象，返回一个嵌套列表，每个元素是一行的cell对象
print(cell_range)
print()
for row in cell_range:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('--- END OF ROW ---')
print()
# 访问特定的行和列
list(sheet.columns)[1] # 获取第二列
list(sheet.rows)[1] # 获取第二行
for cell in list(sheet.columns)[1]:
    print(cell.coordinate, cell.value)
print()

# 写入Excel
wb = openpyxl.Workbook() # 创建一个新的工作簿
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet' # 修改工作表的标题
print(wb.sheetnames)
sheet['A1'] = 'Hello world!'
sheet['A2'] = 'Hello Python!'
sheet['A3'] = 'Hello Excel!'
wb.save('.\Python\Base\example2.xlsx')
print()

# 创建和删除工作表
wb.create_sheet() # 创建一个新的工作表
wb.create_sheet() # 创建一个新的工作表，默认在最后，会自动命名，Sheet2、Sheet3等
wb.create_sheet(index=0, title='First Sheet') # 在指定位置创建一个新的工作表
print(wb.sheetnames)
wb.remove(wb['Sheet']) # 删除一个工作表
del wb['Sheet1'] # 删除一个工作表
print(wb.sheetnames)
# 将值写入单元
sheet['A1'] = 'Hello world!'
sheet['A2'] = 'Hello Python!'

# 设置单元格的字体样式，Font对象
from openpyxl.styles import Font
italic24Font = Font(size=24, italic=True) # italic斜体
sheet['A1'].font = italic24Font
sheet['A1'].value = 'Hello world!'
wb.save('.\Python\Base\example3.xlsx')
## Font对象的属性
# name 字体名称，字符串，字体名称
# size 字体大小，整数
# bold 加粗，布尔值
# italic 斜体，布尔值
# 可以将Font对象保存为变量，然后赋值给单元格的font属性
fontObj1 = Font(name='Times New Roman', bold=True)
print()

# 公式
sheet['A1'] = 200 # 与sheet['A1'].value = 200是一样的，因为value是默认属性
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # 直接在单元格中写入公式

# 设置行高和列宽
sheet.row_dimensions[1].height = 70 # 设置第一行的行高，可以在0-409之间，默认12.75
sheet.column_dimensions['B'].width = 20 # 设置B列的列宽，可以在0-255之间，默认8.43
# 可以将行高和列宽设置为0，这样就隐藏了行或列

# 合并和拆分单元格
# 将一个矩形区域的单元格合并为一个单元格
sheet.merge_cells('A1:D3') # 合并A1到D3的单元格，合并后只有A1有值
# 将一个单元格拆分为一个矩形区域的单元格
sheet.unmerge_cells('A1:D3') # 拆分A1到D3的单元格，拆分后A1到D3都有值，值为A1的值

# 冻结窗口 
# 使得表格的某一行或列不随着用户的滚动而滚动
# 每个Worksheet对象都有一个freeze_panes属性，可以设置冻结窗口
# freeze_panes属性的值是一个字符串，表示冻结窗口的位置，如'A2'表示冻结第一行，'B1'表示冻结第一列
# 'C5'表示冻结第五行和第三列之前的所有行和列，不包括第三列，不包括第五行
# 解冻，将freeze_panes属性设置为None

# 图表
# 创建图表，步骤
# 1. 创建一个Reference对象，表示图表数据的范围
# 2. 创建一个Series对象，表示一个数据系列
# 3. 创建一个Chart对象
# 4. 将Series对象添加到Chart对象
# Reference对象
# Reference对象表示图表数据的范围，可以使用openpyxl.utils中的函数来创建Reference对象
# 创建参数为工作表、行、列、行、列的函数，返回一个Reference对象
# 两个整数的元组表示一个单元格，两个元组的元组表示一个矩形区域，第一个是左上角，第二个是右下角
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
print(refObj)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
print(seriesObj)
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('.\Python\Base\example4.xlsx')
print()

